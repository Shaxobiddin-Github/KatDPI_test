from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.utils import timezone
from .models import Topic, TopicIntroVideo, TopicQuestion, TopicAnswerOption, TopicTest, TopicStudentTest, TopicStudentAnswer, Subject, Group, User
from django.db.models import Sum
import csv, io
try:
    import openpyxl
except ImportError:
    openpyxl = None
from datetime import timedelta
import random
from django.contrib.auth import authenticate, login
import logging

# ---------- Student Simple Login (username only) ----------
def topic_student_login(request):
    """Talaba login: first_name + last_name + group nomi orqali.
    Agar mavjud bo'lmasa yangi student user (random username) yaratiladi.
    Parolsiz (soddalashtirilgan demo kirish)."""
    context = {}
    if request.method == 'POST':
        first_name = (request.POST.get('first_name') or '').strip()
        last_name = (request.POST.get('last_name') or '').strip()
        group_name = (request.POST.get('group') or '').strip()

        # Validation
        if not first_name or not last_name or not group_name:
            context['error'] = 'Barcha maydonlarni to‘ldiring'
        else:
            try:
                grp = Group.objects.get(name__iexact=group_name)
            except Group.DoesNotExist:
                context['error'] = 'Guruh topilmadi'
            else:
                # Try find existing student
                user = User.objects.filter(first_name__iexact=first_name,
                                           last_name__iexact=last_name,
                                           group=grp,
                                           role='student').first()
                if not user:
                    # create new simple user
                    base_username = (first_name + last_name).lower().replace(' ', '')
                    username = base_username
                    counter = 1
                    while User.objects.filter(username=username).exists():
                        counter += 1
                        username = f"{base_username}{counter}"
                    user = User.objects.create(
                        username=username,
                        first_name=first_name,
                        last_name=last_name,
                        role='student',
                        group=grp
                    )
                # Login (no password)
                login(request, user)
                return redirect('topic:student_dashboard')
        context.update({'first_name': first_name, 'last_name': last_name, 'group_name': group_name})
    # Groups list (limit for dropdown)
    context['groups'] = Group.objects.all().order_by('name')[:200]
    return render(request, 'topic_panel/student_login.html', context)

# ---------- Student Dashboard (available tests) ----------
def topic_student_dashboard(request):
    if not request.user.is_authenticated:
        return redirect('topic:student_login')
    if getattr(request.user, 'role', '') != 'student':
        return HttpResponseForbidden('Talaba emas')
    gid = request.user.group_id
    # Tests where this group explicitly assigned
    tests = TopicTest.objects.filter(groups__id=gid, is_published=True).select_related('topic').distinct().order_by('-created_at')
    # Build status info
    runs = {r.test_id: r for r in TopicStudentTest.objects.filter(student=request.user, test_id__in=tests.values_list('id', flat=True))}
    enriched = []
    for t in tests:
        run = runs.get(t.id)
        enriched.append({
            'obj': t,
            'attempt_id': run.id if run else None,
            'completed': run.completed if run else False,
            'score': run.total_score if run else None
        })
    student_full_name = f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username
    group_name = getattr(request.user.group, 'name', None)
    return render(request, 'topic_panel/student_dashboard.html', {
        'tests': enriched,
        'student_full_name': student_full_name,
        'student_group_name': group_name,
    })

# ---------- Helpers ----------

def teacher_required(fn):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated or getattr(request.user, 'role', None) != 'teacher':
            return HttpResponseForbidden('Ruxsat yo\'q')
        return fn(request, *args, **kwargs)
    return wrapper

# ---------- Topic Dashboard ----------
@login_required
@teacher_required
def topic_dashboard(request):
    from .models import Subject
    topics = Topic.objects.filter(created_by=request.user).select_related('subject', 'group')
    subjects = Subject.objects.all().order_by('name')
    # Simple aggregates for header stats
    topic_ids = list(topics.values_list('id', flat=True))
    question_count = 0
    test_count = 0
    if topic_ids:
        question_count = TopicQuestion.objects.filter(topic_id__in=topic_ids).count()
        test_count = TopicTest.objects.filter(topic_id__in=topic_ids).count()
    ctx = {
        'topics': topics,
        'subjects': subjects,
        'teacher_full_name': f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username,
        'teacher_username': request.user.username,
        'agg_questions': question_count,
        'agg_tests': test_count,
        'agg_topics': len(topic_ids),
    }
    return render(request, 'topic_panel/dashboard.html', ctx)

# Safe wrapper for mixed role accidental access
def topic_dashboard_entry(request):
    if not request.user.is_authenticated:
        return redirect('/api/login/?next=/api/topic-panel/dashboard/')
    role = getattr(request.user, 'role', '')
    if role == 'student':
        return redirect('topic:student_dashboard')
    if role == 'teacher':
        return topic_dashboard(request)
    return HttpResponseForbidden('Ruxsat yo\'q')

# ---------- Create Topic ----------
@login_required
@teacher_required
@require_http_methods(["POST"])
@transaction.atomic
def create_topic(request):
    subject_id = request.POST.get('subject_id')
    title = request.POST.get('title')
    description = request.POST.get('description')
    if not (subject_id and title):
        return JsonResponse({'error': 'Fan va sarlavha majburiy'}, status=400)
    # Graceful validation to avoid returning full 404 HTML page
    try:
        subject_id_int = int(subject_id)
    except (TypeError, ValueError):
        return JsonResponse({'error': 'Subject ID noto\'g\'ri'}, status=400)
    subject = Subject.objects.filter(id=subject_id_int).first()
    if not subject:
        return JsonResponse({'error': 'Subject topilmadi'}, status=400)
    topic = Topic.objects.create(subject=subject, title=title, description=description or '', created_by=request.user)
    return JsonResponse({'id': topic.id, 'title': topic.title})

# ---------- Topic Detail (basic JSON) ----------
@login_required
@teacher_required
def topic_detail(request, topic_id):
    topic = get_object_or_404(Topic, id=topic_id, created_by=request.user)
    questions = topic.questions.count()
    tests = topic.tests.count()
    has_video = hasattr(topic, 'intro_video')
    return JsonResponse({'id': topic.id, 'title': topic.title, 'questions': questions, 'tests': tests, 'has_video': has_video})

# ---------- Add Question ----------
@login_required
@teacher_required
@require_http_methods(["POST"])
@transaction.atomic
def add_question(request, topic_id):
    topic = get_object_or_404(Topic, id=topic_id, created_by=request.user)
    text = request.POST.get('text')
    qtype = request.POST.get('question_type')
    if qtype not in ['single', 'multi']:
        return JsonResponse({'error': 'Noto\'g\'ri savol turi'}, status=400)
    if not text:
        return JsonResponse({'error': 'Matn kerak'}, status=400)
    image = request.FILES.get('image')
    question = TopicQuestion.objects.create(topic=topic, text=text, question_type=qtype, image=image, created_by=request.user)
    # Expect options posted as option_text_1..n with is_correct_1..n
    option_keys = [k for k in request.POST.keys() if k.startswith('option_text_')]
    correct_count = 0
    for key in option_keys:
        idx = key.split('_')[-1]
        otext = request.POST.get(key)
        if not otext:
            continue
        is_correct = request.POST.get(f'is_correct_{idx}') == 'on'
        if is_correct:
            correct_count += 1
        TopicAnswerOption.objects.create(question=question, text=otext, is_correct=is_correct)
    # Validation: single => exactly 1 correct; multi => at least 2
    if qtype == 'single' and correct_count != 1:
        question.delete()
        return JsonResponse({'error': 'Bitta to\'g\'ri javob bo\'lishi kerak'}, status=400)
    if qtype == 'multi' and correct_count < 2:
        question.delete()
        return JsonResponse({'error': 'Kamida 2 to\'g\'ri javob bo\'lishi kerak (multi).'}, status=400)
    return JsonResponse({'id': question.id, 'text': question.text})

# ---------- Create Test ----------
@login_required
@teacher_required
@require_http_methods(["POST"])
@transaction.atomic
def create_test(request, topic_id):
    topic = get_object_or_404(Topic, id=topic_id, created_by=request.user)
    title = request.POST.get('title') or f"{topic.title} testi"
    question_count = int(request.POST.get('question_count') or 0)
    total_score = int(request.POST.get('total_score') or 0)
    duration_minutes = int(request.POST.get('duration_minutes') or 0)
    # Optional multiple groups (comma separated IDs or multiple inputs named group_ids)
    raw_group_ids = request.POST.getlist('group_ids') or []
    if len(raw_group_ids) == 1 and ',' in raw_group_ids[0]:
        raw_group_ids = [g.strip() for g in raw_group_ids[0].split(',') if g.strip()]
    if not raw_group_ids:
        return JsonResponse({'error': 'Kamida bitta guruh tanlang'}, status=400)
    if question_count <= 0 or total_score <=0 or duration_minutes <=0:
        return JsonResponse({'error':'Parametrlar to\'liq emas'}, status=400)
    available = list(topic.questions.values_list('id', flat=True))
    if len(available) < question_count:
        return JsonResponse({'error':'Mavzuda yetarli savol yo\'q'}, status=400)
    chosen = random.sample(available, question_count)
    test = TopicTest.objects.create(
        topic=topic,
        title=title,
        question_count=question_count,
        total_score=total_score,
        duration=timedelta(minutes=duration_minutes),
        created_by=request.user,
        question_ids=chosen
    )
    if raw_group_ids:
        valid_groups = Group.objects.filter(id__in=raw_group_ids)
        if not valid_groups.exists():
            return JsonResponse({'error': 'Tanlangan guruh topilmadi'}, status=400)
        test.groups.set(valid_groups)
    return JsonResponse({'id': test.id, 'title': test.title})

# ---------- Start Test (student gating) ----------
@login_required
def start_topic_test(request, test_id):
    test = get_object_or_404(TopicTest, id=test_id)
    if getattr(request.user, 'role', '') != 'student':
        return HttpResponseForbidden('Talaba emas')
    # Group check: ONLY explicitly assigned groups (topic.group not automatic anymore)
    assigned_ids = set(test.groups.values_list('id', flat=True))
    if request.user.group_id not in assigned_ids:
        return HttpResponseForbidden('Bu test siz uchun emas (guruh kiritilmagan)')
    # Pre-video gating
    if hasattr(test.topic, 'intro_video') and not request.session.get(f'topic_video_seen_{test.id}'):
        iv = test.topic.intro_video
        embed_url = None
        if iv.video_url:
            raw = iv.video_url.strip()
            # Convert common YouTube URL formats to embed form
            # Examples handled: https://www.youtube.com/watch?v=ID , https://youtu.be/ID , embed already
            import re
            yt_id = None
            patterns = [
                r'youtube\.com/watch\?v=([A-Za-z0-9_-]{6,})',
                r'youtu\.be/([A-Za-z0-9_-]{6,})',
                r'youtube\.com/embed/([A-Za-z0-9_-]{6,})'
            ]
            for p in patterns:
                m = re.search(p, raw)
                if m:
                    yt_id = m.group(1)
                    break
            if yt_id:
                embed_url = f'https://www.youtube.com/embed/{yt_id}?rel=0&modestbranding=1'
            else:
                # If direct mp4 link or unrecognized, let template try raw in iframe (may fail) or fallback below
                embed_url = raw
        return render(request, 'topic_panel/pre_video.html', {'test': test, 'video': iv, 'embed_url': embed_url})
    # Create or reuse student test
    st, created = TopicStudentTest.objects.get_or_create(student=request.user, test=test)
    if created or not st.randomized_question_ids:
        base_ids = list(test.question_ids)
        # Apply block shuffle if configured
        if test.shuffle_questions and base_ids:
            if test.question_block_size and test.question_block_size > 0:
                block = test.question_block_size
                shuffled = []
                for i in range(0, len(base_ids), block):
                    seg = base_ids[i:i+block]
                    random.shuffle(seg)
                    shuffled.extend(seg)
                base_ids = shuffled
            else:
                random.shuffle(base_ids)
        st.randomized_question_ids = base_ids
        st.save()
        logging.getLogger('api').info('START_TEST_INIT user=%s test=%s created=%s randomized=%s', request.user.id, test.id, created, base_ids)
    else:
        logging.getLogger('api').info('START_TEST_REENTER user=%s test=%s randomized=%s', request.user.id, test.id, st.randomized_question_ids)
    duration_seconds = int(test.duration.total_seconds()) if test.duration else 0
    # Remaining seconds (in case student re-enters)
    remaining = None
    if st.started_at and test.duration:
        elapsed = (timezone.now() - st.started_at).total_seconds()
        remaining = max(0, duration_seconds - int(elapsed))
    return render(request, 'topic_panel/test_run.html', {'test': test, 'student_test': st, 'duration_seconds': remaining if remaining is not None else duration_seconds})

# ---------- Mark video seen ----------
@login_required
def mark_topic_video_seen(request, test_id):
    test = get_object_or_404(TopicTest, id=test_id)
    request.session[f'topic_video_seen_{test.id}'] = True
    return JsonResponse({'ok': True})

# ---------- Submit Answer ----------
@login_required
@require_http_methods(["POST"])
@transaction.atomic
def submit_answer(request, student_test_id):
    st = get_object_or_404(TopicStudentTest, id=student_test_id, student=request.user)
    qid = int(request.POST.get('question_id'))
    q = get_object_or_404(TopicQuestion, id=qid)
    # Enforce time expiry before accepting answers
    test = st.test
    if test.duration and st.started_at:
        elapsed = (timezone.now() - st.started_at).total_seconds()
        if elapsed > test.duration.total_seconds():
            if not st.completed:
                st.completed = True
                st.finished_at = timezone.now()
                st.save(update_fields=['completed', 'finished_at'])
            return JsonResponse({'error': 'Vaqt tugagan'}, status=403)
    selected = request.POST.getlist('options')  # list of option IDs
    ans, _ = TopicStudentAnswer.objects.get_or_create(student_test=st, question=q)
    ans.selected_options.clear()
    chosen_options = TopicAnswerOption.objects.filter(id__in=selected, question=q)
    ans.selected_options.add(*chosen_options)
    # Scoring
    correct_ids = set(q.options.filter(is_correct=True).values_list('id', flat=True))
    chosen_ids = set(chosen_options.values_list('id', flat=True))
    if q.question_type == 'single':
        ans.is_correct = (len(chosen_ids) == 1 and chosen_ids == correct_ids)
    else:  # multi
        ans.is_correct = (chosen_ids == correct_ids)
    # Per-question score = total_score / question_count
    per_q = st.test.total_score / st.test.question_count if st.test.question_count else 0
    ans.score = per_q if ans.is_correct else 0
    ans.save()
    # update total
    agg = st.answers.aggregate(s=Sum('score'))
    st.total_score = agg['s'] or 0
    st.save()
    return JsonResponse({'correct': ans.is_correct, 'score': ans.score, 'total': st.total_score})

# ---------- Finish Test ----------
@login_required
@require_http_methods(["POST"])
@transaction.atomic
def finish_topic_test(request, student_test_id):
    st = get_object_or_404(TopicStudentTest, id=student_test_id, student=request.user)
    # If time expired, still mark completed
    test = st.test
    if test.duration and st.started_at:
        elapsed = (timezone.now() - st.started_at).total_seconds()
        if elapsed > test.duration.total_seconds():
            if not st.completed:
                st.completed = True
                st.finished_at = st.started_at + test.duration
                st.save(update_fields=['completed', 'finished_at'])
    if not st.completed:
        st.completed = True
        st.finished_at = timezone.now()
        st.save(update_fields=['completed', 'finished_at'])
    # Stats
    total_q = len(st.test.question_ids)
    answered = st.answers.count()
    correct = st.answers.filter(is_correct=True).count()
    incorrect = answered - correct
    unanswered = total_q - answered
    return JsonResponse({
        'completed': True,
        'total_score': st.total_score,
        'total_questions': total_q,
        'answered': answered,
        'correct': correct,
        'incorrect': incorrect,
        'unanswered': unanswered
    })

# ---------- Question detail (student fetch) ----------
@login_required
def topic_question_detail(request, question_id):
    q = get_object_or_404(TopicQuestion, id=question_id)
    # permission: ensure student belongs to group when student; teachers can view own topics
    if getattr(request.user, 'role', '') == 'student':
        gid = request.user.group_id
        allowed = False
        # Fetch all runs for this student & group (avoid JSON contains filters unsupported on SQLite)
        candidate_runs = TopicStudentTest.objects.filter(student=request.user, test__groups__id=gid).select_related('test')
        run_ids = []
        for run in candidate_runs:
            run_ids.append(run.id)
            # Accept if question id appears in either stored randomized list or original test question_ids
            rand_list = run.randomized_question_ids or []
            if q.id in rand_list or q.id in (run.test.question_ids or []):
                allowed = True
                logging.getLogger('api').info('QUESTION_ACCESS_OK user=%s q=%s test=%s rand_len=%s', request.user.id, q.id, run.test_id, len(rand_list))
                break
        if not allowed:
            logging.getLogger('api').warning('QUESTION_ACCESS_DENY user=%s q=%s candidate_runs=%s', request.user.id, q.id, run_ids)
            return HttpResponseForbidden("Ruxsat yo'q (savol test ro'yxatida emas yoki guruh mos emas)")
    elif getattr(request.user, 'role', '') == 'teacher':
        if q.topic.created_by_id != request.user.id:
            logging.getLogger('api').warning('QUESTION_ACCESS_DENY_TEACHER user=%s q=%s owner=%s', request.user.id, q.id, q.topic.created_by_id)
            return HttpResponseForbidden('Ruxsat yo\'q')
    # Option shuffling if the owning test has shuffle_options enabled.
    # We detect current active student test for this question (if any) to decide.
    opt_list = list(q.options.all())
    active_run = None  # ensure defined for both roles to avoid UnboundLocalError
    if getattr(request.user, 'role', '') == 'student':
        # Determine active run via Python membership tests (avoid JSON __contains)
        runs = TopicStudentTest.objects.filter(student=request.user).select_related('test').order_by('-started_at')
        for run in runs:
            rand_list = run.randomized_question_ids or []
            if q.id in rand_list or q.id in (run.test.question_ids or []):
                active_run = run
                break
        if active_run and active_run.test.shuffle_options:
            random.shuffle(opt_list)
    logging.getLogger('api').info('QUESTION_DELIVER user=%s q=%s opts=%s active_run=%s shuffle=%s', request.user.id, q.id, [o.id for o in opt_list], active_run.id if active_run else None, bool(active_run and active_run.test.shuffle_options))
    data = {
        'id': q.id,
        'text': q.text,
        'question_type': q.question_type,
        'image': q.image.url if q.image else None,
        'options': [
            {'id': o.id, 'text': o.text}
            for o in opt_list
        ]
    }
    return JsonResponse(data)

# ---------- Simple favicon (inline PNG) ----------
def favicon_view(request):
    # 16x16 transparent PNG (1x1 actually) base64
    import base64
    png_b64 = b'iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR4nGNgYAAAAAMAASsJTYQAAAAASUVORK5CYII='
    data = base64.b64decode(png_b64)
    from django.http import HttpResponse
    return HttpResponse(data, content_type='image/png')

# ---------- Video create/update ----------
@login_required
@teacher_required
@require_http_methods(["POST"])
@transaction.atomic
def save_topic_video(request, topic_id):
    topic = get_object_or_404(Topic, id=topic_id, created_by=request.user)
    video_url = request.POST.get('video_url')
    video_file = request.FILES.get('video_file')
    iv, _ = TopicIntroVideo.objects.get_or_create(topic=topic)
    if video_url:
        iv.video_url = video_url
    if video_file:
        iv.video_file = video_file
    iv.save()
    return JsonResponse({'ok': True, 'video_url': iv.video_url, 'has_file': bool(iv.video_file)})

# ---------- Get video info ----------
@login_required
@teacher_required
def get_topic_video(request, topic_id):
    topic = get_object_or_404(Topic, id=topic_id, created_by=request.user)
    if hasattr(topic, 'intro_video'):
        iv = topic.intro_video
        return JsonResponse({'video_url': iv.video_url, 'video_file': bool(iv.video_file)})
    return JsonResponse({'video_url': None, 'video_file': False})

# ---------- Delete video ----------
@login_required
@teacher_required
@require_http_methods(["POST"])
@transaction.atomic
def delete_topic_video(request, topic_id):
    topic = get_object_or_404(Topic, id=topic_id, created_by=request.user)
    if hasattr(topic, 'intro_video'):
        topic.intro_video.delete()
    return JsonResponse({'ok': True})

# ---------- Manage Topic (teacher UI) ----------
@login_required
@teacher_required
def topic_manage(request, topic_id):
    topic = get_object_or_404(Topic, id=topic_id, created_by=request.user)
    questions = topic.questions.prefetch_related('options').all()
    tests = topic.tests.all().order_by('-created_at')
    video = getattr(topic, 'intro_video', None)
    # Groups for multi-select (teacher may have access restrictions later; for now all groups)
    all_groups = Group.objects.all().order_by('name')
    return render(request, 'topic_panel/topic_manage.html', {
        'topic': topic,
        'questions': questions,
        'tests': tests,
        'video': video,
        'all_groups': all_groups,
    })

# ---------- Test Results (teacher) ----------
@login_required
@teacher_required
def topic_test_results(request, test_id):
    test = get_object_or_404(TopicTest, id=test_id, topic__created_by=request.user)
    # related_name on TopicStudentTest is 'student_runs'
    attempts = test.student_runs.select_related('student').all().order_by('-finished_at','-started_at')
    return render(request, 'topic_panel/test_results.html', {
        'test': test,
        'attempts': attempts
    })

# ---------- Topic Stats (AJAX) ----------
@login_required
@teacher_required
def topic_stats(request, topic_id):
    """Return aggregated statistics for a topic via JSON.
    Counts questions, tests, attempts, and average score across completed attempts.
    """
    topic = get_object_or_404(Topic, id=topic_id, created_by=request.user)
    question_count = topic.questions.count()
    test_count = topic.tests.count()
    runs = TopicStudentTest.objects.filter(test__topic=topic, completed=True)
    attempts = runs.count()
    avg_score = None
    if attempts:
        from django.db.models import Sum
        total_score = runs.aggregate(s=Sum('total_score'))['s'] or 0
        avg_score = round(total_score / attempts, 2)
    return JsonResponse({
        'topic_id': topic.id,
        'title': topic.title,
        'questions': question_count,
        'tests': test_count,
        'attempts': attempts,
        'avg_score': avg_score,
    })

# ---------- Update / Delete Question ----------
@login_required
@teacher_required
@require_http_methods(["POST"])
@transaction.atomic
def update_topic_question(request, topic_id, question_id):
    topic = get_object_or_404(Topic, id=topic_id, created_by=request.user)
    q = get_object_or_404(TopicQuestion, id=question_id, topic=topic)
    text = request.POST.get('text')
    qtype = request.POST.get('question_type')
    if text:
        q.text = text
    if qtype in ['single','multi']:
        q.question_type = qtype
    # Option updates: existing options come as option_<id>=text and correct_<id>=on, plus new_option_text_N
    # Collect existing options
    existing_options = {o.id: o for o in q.options.all()}
    # Track correct count
    correct_count = 0
    for oid, opt in existing_options.items():
        new_text = request.POST.get(f'option_{oid}')
        delete_flag = request.POST.get(f'delete_{oid}')
        if delete_flag == '1':
            opt.delete()
            continue
        if new_text:
            opt.text = new_text
        is_corr = request.POST.get(f'correct_{oid}') == 'on'
        opt.is_correct = is_corr
        if is_corr:
            correct_count += 1
        opt.save()
    # New options pattern: new_option_text_X with new_option_correct_X
    index = 1
    while True:
        ntxt = request.POST.get(f'new_option_text_{index}')
        if not ntxt:
            break
        is_corr = request.POST.get(f'new_option_correct_{index}') == 'on'
        o = TopicAnswerOption.objects.create(question=q, text=ntxt, is_correct=is_corr)
        if is_corr:
            correct_count += 1
        index += 1
    # Re-evaluate correctness constraints
    if q.question_type == 'single':
        # Ensure exactly one correct, else invalidate by picking first correct or set first option correct
        corrects = list(q.options.filter(is_correct=True))
        if len(corrects) != 1 and corrects:
            # Keep only first correct
            for extra in corrects[1:]:
                extra.is_correct = False
                extra.save()
        elif len(corrects) == 0:
            first = q.options.first()
            if first:
                first.is_correct = True
                first.save()
    else:  # multi
        # At least 2 correct; if not, mark first two
        corrects = list(q.options.filter(is_correct=True))
        if len(corrects) < 2:
            all_opts = list(q.options.all())
            for idx,opt in enumerate(all_opts):
                opt.is_correct = (idx < 2)
                opt.save()
    q.save()
    return JsonResponse({'ok': True, 'id': q.id})

@login_required
@teacher_required
@require_http_methods(["POST"])
@transaction.atomic
def delete_topic_question(request, topic_id, question_id):
    topic = get_object_or_404(Topic, id=topic_id, created_by=request.user)
    q = get_object_or_404(TopicQuestion, id=question_id, topic=topic)
    q.delete()
    return JsonResponse({'ok': True})

# ---------- Update / Delete Test ----------
@login_required
@teacher_required
@require_http_methods(["POST"])
@transaction.atomic
def update_topic_test(request, topic_id, test_id):
    topic = get_object_or_404(Topic, id=topic_id, created_by=request.user)
    test = get_object_or_404(TopicTest, id=test_id, topic=topic)
    title = request.POST.get('title')
    total_score = request.POST.get('total_score')
    duration_minutes = request.POST.get('duration_minutes')
    is_published = request.POST.get('is_published')
    new_qcount = request.POST.get('question_count')
    changed = []
    if title:
        test.title = title; changed.append('title')
    if total_score:
        try:
            test.total_score = int(total_score); changed.append('total_score')
        except ValueError:
            pass
    if duration_minutes:
        try:
            from datetime import timedelta
            test.duration = timedelta(minutes=int(duration_minutes)); changed.append('duration')
        except ValueError:
            pass
    if is_published is not None:
        test.is_published = (is_published == '1'); changed.append('is_published')
    # Adjust question set size if requested
    if new_qcount:
        try:
            target = int(new_qcount)
            if target > 0:
                current_ids = list(test.question_ids or [])
                # Fetch all question ids for this topic (teacher's questions for this topic)
                all_ids = list(test.topic.questions.values_list('id', flat=True))
                # If target larger than available pool -> error
                if target > len(all_ids):
                    return JsonResponse({'ok': False, 'error': 'Yetarli savol mavjud emas', 'available': len(all_ids)}, status=400)
                # If expanding: keep existing order, append random remaining distinct ids
                if target > len(current_ids):
                    import random
                    remaining = [i for i in all_ids if i not in current_ids]
                    random.shuffle(remaining)
                    needed = target - len(current_ids)
                    current_ids.extend(remaining[:needed])
                elif target < len(current_ids):
                    current_ids = current_ids[:target]
                if current_ids != (test.question_ids or []):
                    test.question_ids = current_ids
                    test.question_count = len(current_ids)
                    changed.append('question_ids')
        except ValueError:
            pass
    # Optional reassign groups: group_ids repeated
    group_ids = request.POST.getlist('group_ids')
    if group_ids:
        from .models import Group
        groups = Group.objects.filter(id__in=group_ids)
        if groups.exists():
            test.groups.set(groups)
            changed.append('groups')
    test.save()
    return JsonResponse({'ok': True, 'id': test.id, 'changed': changed, 'question_count': test.question_count})

@login_required
@teacher_required
@require_http_methods(["POST"])
@transaction.atomic
def delete_topic_test(request, topic_id, test_id):
    topic = get_object_or_404(Topic, id=topic_id, created_by=request.user)
    test = get_object_or_404(TopicTest, id=test_id, topic=topic)
    # Prevent deleting if students already attempted? (Optional) For now allow.
    test.delete()
    return JsonResponse({'ok': True})

# ---------- Analytics: Group Comparison ----------
@login_required
@teacher_required
def topic_group_comparison(request, topic_id):
    topic = get_object_or_404(Topic, id=topic_id, created_by=request.user)
    # All tests for this topic
    tests = list(topic.tests.all())
    test_ids = [t.id for t in tests]
    # Collect group-level aggregates across all attempts of these tests
    from django.db.models import Count, Avg
    qs = TopicStudentTest.objects.filter(test_id__in=test_ids, completed=True).select_related('student__group')
    data_map = {}
    for run in qs:
        gid = run.student.group_id
        gname = getattr(run.student.group, 'name', '—')
        if gid not in data_map:
            data_map[gid] = {
                'group_name': gname,
                'attempts': 0,
                'total_score': 0.0,
                'max_score': 0.0,
                'students': set(),
            }
        entry = data_map[gid]
        entry['attempts'] += 1
        entry['total_score'] += run.total_score
        entry['max_score'] = max(entry['max_score'], run.total_score)
        entry['students'].add(run.student_id)
    # Post process to compute average
    table = []
    for gid, info in data_map.items():
        avg_score = round(info['total_score']/info['attempts'], 2) if info['attempts'] else 0
        # percent uses max total_score from tests sum for fairness
        max_possible = sum(t.total_score for t in tests) if tests else 0
        avg_percent = round((avg_score / max_possible)*100, 1) if max_possible else 0
        table.append({
            'group_id': gid,
            'group_name': info['group_name'],
            'attempts': info['attempts'],
            'avg_score': avg_score,
            'best_score': info['max_score'],
            'student_count': len(info['students']),
            'avg_percent': avg_percent,
        })
    # Sort by avg_score desc
    table.sort(key=lambda x: x['avg_score'], reverse=True)
    return render(request, 'topic_panel/analytics_group_comparison.html', {
        'topic': topic,
        'rows': table,
        'tests': tests,
    })

# ---------- Analytics: Student Personal Stats ----------
@login_required
def topic_student_personal_stats(request, topic_id):
    # Student only, show own runs per topic
    topic = get_object_or_404(Topic, id=topic_id)
    if getattr(request.user, 'role', '') == 'student':
        runs = TopicStudentTest.objects.filter(student=request.user, test__topic=topic, completed=True).select_related('test').order_by('-finished_at')
        start_date = request.GET.get('start')
        end_date = request.GET.get('end')
        if start_date:
            try:
                from datetime import datetime
                sd = datetime.strptime(start_date, '%Y-%m-%d')
                runs = runs.filter(finished_at__date__gte=sd.date())
            except ValueError:
                pass
        if end_date:
            try:
                from datetime import datetime, timedelta
                ed = datetime.strptime(end_date, '%Y-%m-%d')
                runs = runs.filter(finished_at__date__lte=ed.date())
            except ValueError:
                pass
        items = []
        for r in runs:
            total_q = len(r.test.question_ids)
            correct = r.answers.filter(is_correct=True).count()
            percent = round((correct/total_q)*100, 1) if total_q else 0
            items.append({
                'test_title': r.test.title,
                'score': r.total_score,
                'percent': percent,
                'finished_at': r.finished_at,
                'correct': correct,
                'total_q': total_q,
            })
        # Aggregate statistics for redesigned UI
        if items:
            total_attempts = len(items)
            avg_percent = round(sum(i['percent'] for i in items)/total_attempts, 1)
            best_percent = max(i['percent'] for i in items)
            avg_score = round(sum(i['score'] for i in items)/total_attempts, 1)
            best_score = max(i['score'] for i in items)
            passed = sum(1 for i in items if i['percent'] >= 56)
            pass_rate = round((passed/total_attempts)*100, 1)
            last_date = items[0]['finished_at']  # runs ordered desc
            # Build chronological percent series (oldest -> newest)
            percent_series = [i['percent'] for i in reversed(items)]
            score_series = [i['score'] for i in reversed(items)]
            # Delta vs previous attempt (latest two only)
            delta_percent = 0
            if len(items) >= 2:
                delta_percent = round(items[0]['percent'] - items[1]['percent'], 1)
        else:
            total_attempts = 0
            avg_percent = best_percent = avg_score = best_score = pass_rate = 0
            last_date = None
            percent_series = []
            score_series = []
            delta_percent = 0
        personal_stats = {
            'total_attempts': total_attempts,
            'avg_percent': avg_percent,
            'best_percent': best_percent,
            'avg_score': avg_score,
            'best_score': best_score,
            'pass_rate': pass_rate,
            'last_date': last_date,
            'pass_threshold': 56,
            'percent_series': percent_series,
            'score_series': score_series,
            'delta_percent': delta_percent,
        }
        # Safe JSON for frontend (avoid issues with auto-escaping of complex objects)
        from django.utils.safestring import mark_safe
        import json
        items_js = mark_safe(json.dumps(items, default=str))
        return render(request, 'topic_panel/analytics_student_personal.html', {
            'topic': topic,
            'items': items,
            'items_js': items_js,
            'start_filter': start_date or '',
            'end_filter': end_date or '',
            'personal_stats': personal_stats,
        })
    # Teacher viewing a specific student's stats (optional future). For now forbid.
    return HttpResponseForbidden('Faqat talaba ko‘rishi mumkin')

# ---------- Analytics: Teacher Per-Test Participants Stats ----------
@login_required
@teacher_required
def topic_test_participants_stats(request, topic_id, test_id):
    topic = get_object_or_404(Topic, id=topic_id, created_by=request.user)
    test = get_object_or_404(TopicTest, id=test_id, topic=topic)
    runs = TopicStudentTest.objects.filter(test=test, completed=True)\
        .select_related('student')\
        .prefetch_related('answers__question', 'answers__selected_options', 'answers')

    rows = []  # per-student summary

    # --- Build per-question matrix data ---
    question_id_order = test.question_ids or []
    # Fetch all involved questions & options to avoid N+1
    from collections import defaultdict
    q_objs = {q.id: q for q in TopicQuestion.objects.filter(id__in=question_id_order).prefetch_related('options')}

    # Prepare question meta list preserving order
    question_meta = []  # [{id, text_short, correct_option_ids, option_letters:{opt_id:'A'}, is_multi}]
    option_letter_map = {}  # opt_id -> letter
    for qid in question_id_order:
        q = q_objs.get(qid)
        if not q:
            continue
        letters = {}
        letter_seq = []
        # Assign letters A, B, C ... based on current ordering of options queryset
        for idx, opt in enumerate(q.options.all()):
            letter = chr(ord('A') + idx)
            letters[opt.id] = letter
            option_letter_map[opt.id] = letter
            letter_seq.append((opt.id, letter))
        correct_ids = [opt.id for opt in q.options.all() if opt.is_correct]
        question_meta.append({
            'id': q.id,
            'text_short': (q.text[:40] + '...') if len(q.text) > 43 else q.text,
            'correct_option_ids': correct_ids,
            'is_multi': q.question_type.lower() in ['multi', 'multiple', 'multiple_choice', 'multi_choice'],
            'letters': letters,
            'option_sequence': letter_seq,
            'option_count': len(letter_seq),
        })

    # Per-question aggregate counters
    per_question_correct_counts = defaultdict(int)
    per_question_answered_counts = defaultdict(int)

    # Student answer matrix: list aligned with rows after sorting later
    # We'll first collect unsorted then re-order after computing percent
    matrix_entries = []  # each element: {'run_id':..., 'answers': {question_id: {'selected':[opt_ids], 'is_correct':bool}}}

    for run in runs:
        answers = run.answers.all()
        # Map question_id -> answer object
        ans_map = {a.question_id: a for a in answers}
        correct = 0
        incorrect = 0
        student_answers_struct = {}
        for qmeta in question_meta:
            qid = qmeta['id']
            aobj = ans_map.get(qid)
            if not aobj:
                continue  # Unanswered (skipped)
            sel_ids = list(aobj.selected_options.values_list('id', flat=True))
            if aobj.is_correct:
                correct += 1
                per_question_correct_counts[qid] += 1
            else:
                incorrect += 1
            per_question_answered_counts[qid] += 1
            student_answers_struct[qid] = {
                'selected_option_ids': sel_ids,
                'is_correct': aobj.is_correct,
            }
        total_q = len(question_meta) or (correct + incorrect)
        percent = round((correct/total_q)*100, 1) if total_q else 0
        rows.append({
            'student': run.student,
            'student_name': f"{run.student.first_name} {run.student.last_name}".strip() or run.student.username,
            'group_name': getattr(run.student.group, 'name', ''),
            'correct': correct,
            'incorrect': incorrect,
            'score': run.total_score,
            'percent': percent,
            'finished_at': run.finished_at,
            'run_id': run.id,
        })
        matrix_entries.append({
            'run_id': run.id,
            'answers': student_answers_struct,
        })

    # Sort students by percent desc (stable)
    rows.sort(key=lambda r: r['percent'], reverse=True)
    # Reindex matrix_entries to sorted order
    runid_to_matrix = {m['run_id']: m for m in matrix_entries}
    # Attach answers directly to each row for template access
    for r in rows:
        m = runid_to_matrix.get(r['run_id'])
        r['answers'] = (m or {}).get('answers', {})
    # Keep matrix_ordered if needed later (legacy, may remove)
    matrix_ordered = [runid_to_matrix[r['run_id']] for r in rows]

    # Build per-question summary (percent correct)
    question_summaries = []
    for qmeta in question_meta:
        qid = qmeta['id']
        ans_cnt = per_question_answered_counts[qid]
        corr_cnt = per_question_correct_counts[qid]
        percent_corr = round((corr_cnt/ans_cnt)*100, 1) if ans_cnt else 0
        question_summaries.append({
            'question_id': qid,
            'answered': ans_cnt,
            'correct': corr_cnt,
            'percent_correct': percent_corr,
        })

    # Performance safeguard flags
    too_many = False
    MAX_Q = 120
    MAX_STUDENTS = 250
    if len(question_meta) > MAX_Q or len(rows) > MAX_STUDENTS:
        too_many = True

    # Calculate additional statistics for the template
    stats = {
        'total_participants': len(rows),
        'passed_count': sum(1 for r in rows if r['percent'] >= 56),
        'average_percent': round(sum(r['percent'] for r in rows) / len(rows), 1) if rows else 0,
        'average_score': round(sum(r['score'] for r in rows) / len(rows), 1) if rows else 0,
        'total_questions': len(question_meta),
    }
    stats['pass_rate'] = round((stats['passed_count'] / stats['total_participants']) * 100, 1) if stats['total_participants'] else 0

    return render(request, 'topic_panel/analytics_test_participants.html', {
        'topic': topic,
        'test': test,
        'rows': rows,
        'question_meta': question_meta,
        'matrix_entries': matrix_ordered,
        'question_summaries': question_summaries,
        'too_many': too_many,
        'stats': stats,
    })

# ---------- Import Questions (CSV/XLSX) ----------
@login_required
@teacher_required
def topic_import_questions(request, topic_id):
    topic = get_object_or_404(Topic, id=topic_id, created_by=request.user)
    if request.method == 'GET':
        return render(request, 'topic_panel/import_questions.html', {'topic': topic, 'has_openpyxl': openpyxl is not None})
    # POST
    f = request.FILES.get('file')
    if not f:
        return JsonResponse({'error': 'Fayl topilmadi'}, status=400)
    name = f.name.lower()
    rows = []
    errors = []
    created = 0
    # Expected columns
    # text,question_type,option1,correct1,option2,correct2,... up to option6
    try:
        if name.endswith('.csv'):
            data = f.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(data))
            rows = list(reader)
        elif (name.endswith('.xlsx') or name.endswith('.xls')) and openpyxl:
            wb = openpyxl.load_workbook(f)
            ws = wb.active
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for r in ws.iter_rows(min_row=2):
                d = {}
                for idx, cell in enumerate(r):
                    key = headers[idx] if idx < len(headers) else f'col{idx}'
                    d[key] = cell.value
                rows.append(d)
        else:
            return JsonResponse({'error': 'Faqat CSV yoki XLSX (openpyxl o\'rnatilgan bo\'lsa)'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Faylni o\'qishda xatolik: {e}'}, status=400)

    for idx, r in enumerate(rows, start=2):  # start=2 for header offset
        text = (r.get('text') or '').strip()
        qtype = (r.get('question_type') or 'single').strip().lower()
        if qtype not in ['single', 'multi']:
            errors.append(f"{idx}-qator: question_type noto'g'ri")
            continue
        if not text:
            errors.append(f"{idx}-qator: text bo'sh")
            continue
        # Collect options
        options = []
        correct_count = 0
        for i in range(1, 7):
            otext = r.get(f'option{i}')
            if not otext:
                continue
            is_corr_raw = str(r.get(f'correct{i}') or '').strip().lower()
            is_corr = is_corr_raw in ['1','true','yes','ha','y']
            if is_corr:
                correct_count += 1
            options.append((otext, is_corr))
        if len(options) < 2:
            errors.append(f"{idx}-qator: kamida 2 ta variant kerak")
            continue
        if qtype == 'single' and correct_count != 1:
            errors.append(f"{idx}-qator: single uchun aniq 1 ta to'g'ri kerak")
            continue
        if qtype == 'multi' and correct_count < 2:
            errors.append(f"{idx}-qator: multi uchun kamida 2 ta to'g'ri kerak")
            continue
        # Create question + options
        q = TopicQuestion.objects.create(topic=topic, text=text, question_type=qtype, created_by=request.user)
        for otext, isc in options:
            TopicAnswerOption.objects.create(question=q, text=otext, is_correct=isc)
        created += 1
    return JsonResponse({'created': created, 'errors': errors, 'total_rows': len(rows)})

# ---------- Export Questions (CSV) ----------
@login_required
@teacher_required
def topic_export_questions(request, topic_id):
    topic = get_object_or_404(Topic, id=topic_id, created_by=request.user)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['text','question_type','option1','correct1','option2','correct2','option3','correct3','option4','correct4','option5','correct5','option6','correct6'])
    for q in topic.questions.prefetch_related('options').all():
        row = [q.text, q.question_type]
        opts = list(q.options.all())[:6]
        for o in opts:
            row.append(o.text)
            row.append('1' if o.is_correct else '0')
        # pad
        while len(row) < 14:  # 2 + (6*2)=14
            row.append('')
        writer.writerow(row)
    resp = HttpResponse(output.getvalue(), content_type='text/csv')
    resp['Content-Disposition'] = f'attachment; filename=topic_{topic.id}_questions.csv'
    return resp

@login_required
@teacher_required
def topic_export_questions_xlsx(request, topic_id):
    if openpyxl is None:
        return JsonResponse({'error': 'openpyxl o\'rnatilmagan'}, status=400)
    topic = get_object_or_404(Topic, id=topic_id, created_by=request.user)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Savollar'
    headers = ['text','question_type','option1','correct1','option2','correct2','option3','correct3','option4','correct4','option5','correct5','option6','correct6']
    ws.append(headers)
    for q in topic.questions.prefetch_related('options').all():
        row = [q.text, q.question_type]
        opts = list(q.options.all())[:6]
        for o in opts:
            row.append(o.text)
            row.append(1 if o.is_correct else 0)
        while len(row) < len(headers):
            row.append('')
        ws.append(row)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename=topic_{topic.id}_questions.xlsx'
    return resp

# ---------- Remaining Time (AJAX) ----------
@login_required
def test_remaining_time(request, student_test_id):
    st = get_object_or_404(TopicStudentTest, id=student_test_id, student=request.user)
    test = st.test
    remaining = None
    if test.duration and st.started_at:
        total = test.duration.total_seconds()
        elapsed = (timezone.now() - st.started_at).total_seconds()
        remaining = max(0, int(total - elapsed))
        if remaining == 0 and not st.completed:
            st.completed = True
            st.finished_at = st.started_at + test.duration
            st.save(update_fields=['completed', 'finished_at'])
    return JsonResponse({'remaining_seconds': remaining})
